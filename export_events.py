import openpyxl
from openpyxl.utils import get_column_letter
from datetime import timedelta, date, datetime

def generate_event_dates(start, end, recurrence, holidays):
    dates = []

    # 🔹 Convert strings to date objects if necessary
    if isinstance(start, str):
        start = datetime.strptime(start, "%Y-%m-%d").date()
    if isinstance(end, str):
        end = datetime.strptime(end, "%Y-%m-%d").date()

    # 🔹 Ensure holidays are all date objects
    holidays = {h.date() if isinstance(h, datetime) else h for h in holidays}

    current = start
    while current <= end:
        if current not in holidays:
            dates.append(current.strftime('%Y-%m-%d'))

        # 🔄 Recurrence stepping
        if recurrence == 'daily':
            current += timedelta(days=1)
        elif recurrence == 'weekly':
            current += timedelta(weeks=1)
        elif recurrence == 'monthly':
            # 🧠 Handle edge case like Jan 31 -> Feb 28
            next_month = current.month + 1 if current.month < 12 else 1
            next_year = current.year if current.month < 12 else current.year + 1
            try:
                current = current.replace(year=next_year, month=next_month)
            except ValueError:
                # Fallback to last day of month
                current = (current.replace(day=1, year=next_year, month=next_month) + timedelta(days=31)).replace(day=1) - timedelta(days=1)
        else:
            # One-time event — break immediately
            break

    return dates


# def export_to_excel(events, holidays, file_path):
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "User Events"

#     headers = ["Event ID", "User ID", "Event Name", "Event Date"]
#     ws.append(headers)

#     for event in events:
#         event_id, user_id, name, start, end, recurrence = event
#         all_dates = generate_event_dates(start, end, recurrence, holidays)

#         for date in all_dates:
#             ws.append([event_id, user_id, name, date])

#     for col in ws.columns:
#         max_len = max(len(str(cell.value or "")) for cell in col)
#         ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

#     wb.save(file_path)

def export_to_excel(events, holidays, file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Events"

    headers = ["Event ID", "User ID", "Event Name", "Event Date"]
    ws.append(headers)

    for idx, event in enumerate(events):
        print(f"Row {idx} ➜ {event} (length: {len(event)})")  # 🔍 DEBUG

        try:
            event_id, user_id, name, start, end, recurrence = event
        except Exception as e:
            print(f"❌ Error unpacking row {idx}: {event} — {e}")
            continue  # Skip this row

        all_dates = generate_event_dates(start, end, recurrence, holidays)

        for date in all_dates:
            ws.append([event_id, user_id, name, date])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    wb.save(file_path)
